from django.shortcuts import render, get_object_or_404, redirect
from django.http import JsonResponse, HttpResponse
from django.views.decorators.http import require_POST
from django.utils import timezone
from .models import Docente, Estudiante, Materia, Curso, Calificacion, Tarea, Asistencia, Mensaje
from datetime import datetime

DOCENTE_ID = 1

def get_docente():
    return Docente.objects.filter(id=DOCENTE_ID).first()

def dashboard_docente(request):
    docente = get_docente()
    if not docente:
        return render(request, 'calificaciones/dashboard_docente.html', {})
    cursos = docente.cursos.all()
    total_estudiantes = Estudiante.objects.filter(curso__in=cursos).count()
    total_tareas = Tarea.objects.filter(docente=docente).count()
    calificaciones = Calificacion.objects.filter(docente=docente)
    promedio = round(sum(c.promedio for c in calificaciones) / calificaciones.count(), 1) if calificaciones.exists() else 0
    context = {
        'docente': docente, 'cursos': cursos,
        'total_estudiantes': total_estudiantes,
        'total_tareas': total_tareas, 'promedio': promedio,
        'tareas_recientes': Tarea.objects.filter(docente=docente).order_by('-fecha_limite')[:3],
    }
    return render(request, 'calificaciones/dashboard_docente.html', context)

def cursos_docente(request):
    docente = get_docente()
    cursos = docente.cursos.all() if docente else []
    context = {
        'docente': docente, 'cursos': cursos,
        'total_estudiantes': Estudiante.objects.filter(curso__in=cursos).count() if docente else 0,
        'total_materias': docente.materias.count() if docente else 0,
    }
    return render(request, 'calificaciones/cursos_docente.html', context)

def tarea_docente(request):
    docente = get_docente()
    tareas = Tarea.objects.filter(docente=docente) if docente else []
    cursos = docente.cursos.all() if docente else []
    materias = docente.materias.all() if docente else []
    if request.method == 'POST':
        titulo = request.POST.get('titulo')
        descripcion = request.POST.get('descripcion', '')
        materia_id = request.POST.get('materia')
        curso_id = request.POST.get('curso')
        fecha_limite = request.POST.get('fecha_limite')
        if titulo and materia_id and curso_id and fecha_limite:
            Tarea.objects.create(
                titulo=titulo, descripcion=descripcion,
                materia_id=materia_id, curso_id=curso_id,
                docente=docente, fecha_limite=fecha_limite,
                estado='pendiente'
            )
        return redirect('tarea_docente')
    context = {
        'docente': docente, 'tareas': tareas, 'cursos': cursos, 'materias': materias,
        'pendientes': tareas.filter(estado='pendiente').count() if hasattr(tareas, 'filter') else 0,
        'entregadas': tareas.filter(estado='entregada').count() if hasattr(tareas, 'filter') else 0,
        'total_tareas': tareas.count() if hasattr(tareas, 'count') else 0,
    }
    return render(request, 'calificaciones/tarea_docente.html', context)

def eliminar_tarea(request, id):
    tarea = get_object_or_404(Tarea, id=id)
    tarea.delete()
    return redirect('tarea_docente')

def calificaciones_docente(request):
    docente = get_docente()
    cursos = docente.cursos.all() if docente else []
    materias = docente.materias.all() if docente else []
    curso_id = request.GET.get('curso')
    materia_id = request.GET.get('materia')
    curso_sel = None
    materia_sel = None
    estudiantes = []
    calificaciones_map = {}
    if curso_id and materia_id:
        curso_sel = get_object_or_404(Curso, id=curso_id)
        materia_sel = get_object_or_404(Materia, id=materia_id)
        estudiantes = Estudiante.objects.filter(curso=curso_sel)
        for est in estudiantes:
            cal = Calificacion.objects.filter(
                estudiante=est, materia=materia_sel, docente=docente
            ).first()
            calificaciones_map[est.id] = cal
    total_cals = Calificacion.objects.filter(docente=docente)
    promedio_general = round(
        sum(c.promedio for c in total_cals) / total_cals.count(), 1
    ) if total_cals.exists() else 0
    context = {
        'docente': docente, 'cursos': cursos, 'materias': materias,
        'curso_sel': curso_sel, 'materia_sel': materia_sel,
        'estudiantes': estudiantes, 'calificaciones_map': calificaciones_map,
        'promedio_general': promedio_general,
        'en_riesgo':   sum(1 for c in total_cals if c.promedio < 3.0),
        'destacados':  sum(1 for c in total_cals if c.promedio >= 4.5),
        'total_notas': total_cals.count(),
    }
    return render(request, 'calificaciones/calificaciones_docente.html', context)


@require_POST
def guardar_nota(request):
    try:
        docente       = get_docente()
        estudiante_id = request.POST.get('estudiante_id')
        materia_id    = request.POST.get('materia_id')
        tarea         = round(float(request.POST.get('tarea',   0)), 1)
        parcial       = round(float(request.POST.get('parcial', 0)), 1)
        examen        = round(float(request.POST.get('examen',  0)), 1)
        promedio      = round((tarea + parcial + examen) / 3, 2)

        estudiante = get_object_or_404(Estudiante, id=estudiante_id)

        Calificacion.objects.update_or_create(
            estudiante=estudiante,
            materia_id=materia_id,
            defaults={
                'docente': docente,
                'tarea':   tarea,
                'parcial': parcial,
                'examen':  examen,
            }
        )
        return JsonResponse({'ok': True, 'promedio': str(promedio)})

    except Exception as e:
        import traceback
        traceback.print_exc()
        return JsonResponse({'ok': False, 'error': str(e)}, status=400)

@require_POST
def eliminar_nota(request):
    try:
        estudiante_id = request.POST.get('estudiante_id')
        materia_id    = request.POST.get('materia_id')
        Calificacion.objects.filter(
            estudiante_id=estudiante_id,
            materia_id=materia_id,
            docente=get_docente()
        ).delete()
        return JsonResponse({'ok': True})
    except Exception as e:
        return JsonResponse({'ok': False, 'error': str(e)}, status=400)


@require_POST
def guardar_asistencia(request):
    try:
        estudiante_id = request.POST.get('estudiante_id')
        fecha_str     = request.POST.get('fecha')
        estado        = request.POST.get('estado', 'presente')

        estudiante = get_object_or_404(Estudiante, id=estudiante_id)
        fecha      = datetime.strptime(fecha_str, '%Y-%m-%d').date()

        Asistencia.objects.update_or_create(
            estudiante=estudiante,
            fecha=fecha,
            defaults={'estado': estado}
        )
        return JsonResponse({'ok': True})

    except Exception as e:
        import traceback
        traceback.print_exc()
        return JsonResponse({'ok': False, 'error': str(e)}, status=400)


def reporte_notas(request):
    docente  = get_docente()
    cursos   = docente.cursos.all() if docente else []
    materias = docente.materias.all() if docente else []

    curso_id   = request.GET.get('curso')
    materia_id = request.GET.get('materia')
    periodo    = request.GET.get('periodo')

    curso_sel   = None
    materia_sel = None

    calificaciones = Calificacion.objects.filter(docente=docente)

    if curso_id:
        curso_sel      = get_object_or_404(Curso, id=curso_id)
        calificaciones = calificaciones.filter(estudiante__curso=curso_sel)

    if materia_id:
        materia_sel    = get_object_or_404(Materia, id=materia_id)
        calificaciones = calificaciones.filter(materia=materia_sel)

    calificaciones = calificaciones.select_related('estudiante', 'estudiante__curso', 'materia')

    total_estudiantes = calificaciones.count()
    promedio_grupo = round(
        sum(c.promedio for c in calificaciones) / total_estudiantes, 1
    ) if total_estudiantes else 0
    en_riesgo = sum(1 for c in calificaciones if c.promedio < 3.0)
    aprobados  = sum(1 for c in calificaciones if c.promedio >= 3.0)

    context = {
        'docente': docente, 'cursos': cursos, 'materias': materias,
        'curso_sel': curso_sel, 'materia_sel': materia_sel,
        'periodo_sel': periodo,
        'calificaciones': calificaciones,
        'total_estudiantes': total_estudiantes,
        'promedio_grupo': promedio_grupo,
        'en_riesgo': en_riesgo,
        'aprobados': aprobados,
    }
    return render(request, 'calificaciones/reporte_notas.html', context)


def exportar_reporte_pdf(request):
    from reportlab.lib.pagesizes import landscape, letter
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
    from reportlab.lib.styles import getSampleStyleSheet
    from reportlab.lib import colors

    docente    = get_docente()
    curso_id   = request.GET.get('curso')
    materia_id = request.GET.get('materia')

    calificaciones = Calificacion.objects.filter(docente=docente)
    if curso_id:
        calificaciones = calificaciones.filter(estudiante__curso_id=curso_id)
    if materia_id:
        calificaciones = calificaciones.filter(materia_id=materia_id)
    calificaciones = calificaciones.select_related('estudiante', 'estudiante__curso', 'materia')

    response = HttpResponse(content_type='application/pdf')
    response['Content-Disposition'] = 'attachment; filename="reporte_notas.pdf"'

    doc = SimpleDocTemplate(response, pagesize=landscape(letter), topMargin=30, bottomMargin=30)
    styles = getSampleStyleSheet()
    elements = []

    elements.append(Paragraph("Reporte de Notas", styles['Title']))
    if docente:
        elements.append(Paragraph(f"Docente: {docente.nombre}", styles['Normal']))
    elements.append(Spacer(1, 12))

    def desempeno(p):
        if p >= 4.5: return 'Superior'
        if p >= 4.0: return 'Alto'
        if p >= 3.0: return 'Básico'
        return 'Bajo'

    headers = ['#', 'Estudiante', 'Curso', 'Materia', 'Tarea', 'Parcial', 'Examen', 'Promedio', 'Desempeño']
    data = [headers]
    for i, cal in enumerate(calificaciones, 1):
        data.append([
            str(i), cal.estudiante.nombre, cal.estudiante.curso.nombre,
            cal.materia.nombre, str(cal.tarea), str(cal.parcial),
            str(cal.examen), str(cal.promedio), desempeno(cal.promedio),
        ])

    tabla = Table(data, repeatRows=1)
    tabla.setStyle(TableStyle([
        ('BACKGROUND',     (0, 0), (-1, 0), colors.HexColor('#006eff')),
        ('TEXTCOLOR',      (0, 0), (-1, 0), colors.white),
        ('FONTNAME',       (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE',       (0, 0), (-1, 0), 10),
        ('ALIGN',          (0, 0), (-1,-1), 'CENTER'),
        ('FONTSIZE',       (0, 1), (-1,-1), 9),
        ('ROWBACKGROUNDS', (0, 1), (-1,-1), [colors.white, colors.HexColor('#f8fafc')]),
        ('GRID',           (0, 0), (-1,-1), 0.4, colors.HexColor('#e5e7eb')),
        ('TOPPADDING',     (0, 0), (-1,-1), 6),
        ('BOTTOMPADDING',  (0, 0), (-1,-1), 6),
    ]))
    elements.append(tabla)
    doc.build(elements)
    return response


def exportar_reporte_excel(request):
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

    docente    = get_docente()
    curso_id   = request.GET.get('curso')
    materia_id = request.GET.get('materia')

    calificaciones = Calificacion.objects.filter(docente=docente)
    if curso_id:
        calificaciones = calificaciones.filter(estudiante__curso_id=curso_id)
    if materia_id:
        calificaciones = calificaciones.filter(materia_id=materia_id)
    calificaciones = calificaciones.select_related('estudiante', 'estudiante__curso', 'materia')

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Reporte de Notas'

    ws.merge_cells('A1:I1')
    ws['A1'] = 'REPORTE DE NOTAS'
    ws['A1'].font = Font(bold=True, size=14, color='FFFFFF')
    ws['A1'].fill = PatternFill('solid', fgColor='006EFF')
    ws['A1'].alignment = Alignment(horizontal='center')

    if docente:
        ws.merge_cells('A2:I2')
        ws['A2'] = f'Docente: {docente.nombre}'
        ws['A2'].alignment = Alignment(horizontal='center')
        ws['A2'].font = Font(italic=True, size=10)

    thin = Border(
        left=Side(style='thin', color='E5E7EB'),
        right=Side(style='thin', color='E5E7EB'),
        top=Side(style='thin', color='E5E7EB'),
        bottom=Side(style='thin', color='E5E7EB'),
    )
    headers = ['#', 'Estudiante', 'Curso', 'Materia', 'Tarea', 'Parcial', 'Examen', 'Promedio', 'Desempeño']
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=4, column=col, value=h)
        cell.fill = PatternFill('solid', fgColor='081B3A')
        cell.font = Font(bold=True, color='FFFFFF', size=10)
        cell.alignment = Alignment(horizontal='center')
        cell.border = thin

    def desempeno(p):
        if p >= 4.5: return 'Superior'
        if p >= 4.0: return 'Alto'
        if p >= 3.0: return 'Básico'
        return 'Bajo'

    color_desempeno = {'Superior': 'DBEAFE', 'Alto': 'D1FAE5', 'Básico': 'FEF9C3', 'Bajo': 'FEE2E2'}

    for i, cal in enumerate(calificaciones, 1):
        row = i + 4
        d   = desempeno(cal.promedio)
        bg  = 'FFFFFF' if i % 2 == 0 else 'F8FAFC'
        fila = [i, cal.estudiante.nombre, cal.estudiante.curso.nombre,
                cal.materia.nombre, float(cal.tarea), float(cal.parcial),
                float(cal.examen), float(cal.promedio), d]
        for col, val in enumerate(fila, 1):
            cell = ws.cell(row=row, column=col, value=val)
            cell.border = thin
            cell.alignment = Alignment(horizontal='center')
            cell.fill = PatternFill('solid', fgColor=bg)
            if col == 9:
                cell.fill = PatternFill('solid', fgColor=color_desempeno.get(d, 'FFFFFF'))
            if col == 8:
                cell.font = Font(bold=True, color='10B981' if cal.promedio >= 3.0 else 'EF4444')

    anchos = [5, 30, 15, 20, 10, 10, 10, 12, 14]
    for col, ancho in enumerate(anchos, 1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(col)].width = ancho

    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = 'attachment; filename="reporte_notas.xlsx"'
    wb.save(response)
    return response


def asistencia_docente(request):
    docente = get_docente()
    cursos = docente.cursos.all() if docente else []
    curso_id = request.GET.get('curso')
    curso_sel = None
    estudiantes = []
    if curso_id:
        curso_sel = get_object_or_404(Curso, id=curso_id)
        estudiantes = Estudiante.objects.filter(curso=curso_sel)
    asistencias_map = {}
    for est in estudiantes:
        asis = Asistencia.objects.filter(
            estudiante=est, fecha=timezone.now().date()
        ).first()
        asistencias_map[est.id] = asis.estado if asis else 'presente'
    context = {
        'docente': docente, 'cursos': cursos, 'curso_sel': curso_sel,
        'estudiantes': estudiantes, 'asistencias_map': asistencias_map,
        'fecha_hoy': timezone.now().date(),
        'presentes':  sum(1 for v in asistencias_map.values() if v == 'presente'),
        'ausentes':   sum(1 for v in asistencias_map.values() if v == 'ausente'),
        'tardanzas':  sum(1 for v in asistencias_map.values() if v == 'tardanza'),
        'total': len(estudiantes),
    }
    return render(request, 'calificaciones/asistencia_docente.html', context)


def mensajes_docente(request):
    docente = get_docente()
    estudiantes = Estudiante.objects.filter(curso__in=docente.cursos.all()) if docente else []
    estudiante_id = request.GET.get('estudiante')
    estudiante_sel = None
    mensajes = []
    if estudiante_id:
        estudiante_sel = get_object_or_404(Estudiante, id=estudiante_id)
        mensajes = Mensaje.objects.filter(docente=docente, estudiante=estudiante_sel)
        Mensaje.objects.filter(
            docente=docente, estudiante=estudiante_sel, enviado_por='estudiante'
        ).update(leido=True)
    if request.method == 'POST':
        contenido = request.POST.get('contenido')
        est_id = request.POST.get('estudiante_id')
        if contenido and est_id:
            Mensaje.objects.create(
                docente=docente, estudiante_id=est_id,
                contenido=contenido, enviado_por='docente'
            )
        return redirect(f'/docente/mensajes/?estudiante={est_id}')
    sin_leer = Mensaje.objects.filter(
        docente=docente, enviado_por='estudiante', leido=False
    ).count() if docente else 0
    context = {
        'docente': docente, 'estudiantes': estudiantes,
        'estudiante_sel': estudiante_sel, 'mensajes': mensajes,
        'sin_leer': sin_leer,
        'total_mensajes': Mensaje.objects.filter(docente=docente).count() if docente else 0,
    }
    return render(request, 'calificaciones/mensajes_docente.html', context)


def configuracion_docente(request):
    docente = get_docente()
    if request.method == 'POST' and docente:
        docente.nombre   = request.POST.get('nombre',   docente.nombre)
        docente.correo   = request.POST.get('correo',   docente.correo)
        docente.telefono = request.POST.get('telefono', docente.telefono)
        docente.save()
    return render(request, 'calificaciones/configuracion_docente.html', {'docente': docente})